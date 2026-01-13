"""
Excel 文件处理器模块
支持 .xls 和 .xlsx 格式
"""

import xlrd
import openpyxl


class XlsHandler:
    """处理旧版 .xls 格式文件"""
    
    def __init__(self, filepath):
        self.book = xlrd.open_workbook(filepath, formatting_info=True)
        self.sheet = self.book.sheet_by_index(0)
        self.nrows = self.sheet.nrows
        self.ncols = self.sheet.ncols

    def get_cell_value(self, r, c):
        return self.sheet.cell_value(r, c) if r < self.nrows and c < self.ncols else ""

    def is_blue_cell(self, r, c):
        if r >= self.nrows or c >= self.ncols:
            return False
        xf_idx = self.sheet.cell_xf_index(r, c)
        return self.book.xf_list[xf_idx].background.pattern_colour_index == 31


class XlsxHandler:
    """处理新版 .xlsx 格式文件"""
    
    def __init__(self, filepath):
        self.wb = openpyxl.load_workbook(filepath, data_only=True)
        self.sheet = self.wb.active
        self.nrows = self.sheet.max_row
        self.ncols = self.sheet.max_column

    def get_cell_value(self, r, c):
        try:
            return self.sheet.cell(row=r + 1, column=c + 1).value
        except:
            return ""

    def is_blue_cell(self, r, c):
        try:
            cell = self.sheet.cell(row=r + 1, column=c + 1)
            return cell.fill.fgColor.type == "theme" and cell.fill.fgColor.theme == 4
        except:
            return False


def get_handler(filepath):
    """根据文件扩展名返回对应的处理器"""
    if filepath.lower().endswith(".xls"):
        return XlsHandler(filepath)
    return XlsxHandler(filepath)
